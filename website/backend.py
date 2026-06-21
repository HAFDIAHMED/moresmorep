"""Reservation backend for moresolustions-moreproblems.com.

Tiny Flask service that:
  - GET  /status   → how many signed copies are still available
  - POST /reserve  → record a reservation (name + address + optional location)
                     into reservations.xlsx, capped at LIMIT entries.

Runs on 127.0.0.1:5051. Nginx proxies /api/* to this service.
Excel file lives alongside this script (same folder as index.html on the server)."""

from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from datetime import datetime, timezone
from pathlib import Path
from threading import Lock

HERE = Path(__file__).resolve().parent
XLSX = HERE / 'reservations.xlsx'
LIMIT = 10
HEADERS = ['Timestamp (UTC)', 'Position', 'Name', 'Address', 'Location', 'IP', 'Latitude', 'Longitude']

app = Flask(__name__)
_lock = Lock()


def ensure_workbook():
    """Create or migrate the Excel file so its header row matches HEADERS.
    Existing data rows keep their column positions; new columns at the end
    just stay empty for pre-existing rows."""
    if not XLSX.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reservations'
        ws.append(HEADERS)
        widths = {'A': 22, 'B': 10, 'C': 28, 'D': 60, 'E': 28, 'F': 18, 'G': 14, 'H': 14}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        wb.save(XLSX)
        return

    # File exists — extend the header row in place if the schema has grown
    wb = load_workbook(XLSX)
    ws = wb.active
    current = [c.value for c in ws[1]]
    if current[:len(HEADERS)] != HEADERS:
        for idx, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=idx, value=h)
        widths = {'A': 22, 'B': 10, 'C': 28, 'D': 60, 'E': 28, 'F': 18, 'G': 14, 'H': 14}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        wb.save(XLSX)
    wb.close()


def count_rows():
    ensure_workbook()
    wb = load_workbook(XLSX, read_only=True)
    n = wb.active.max_row - 1
    wb.close()
    return max(n, 0)


def client_ip():
    fwd = request.headers.get('X-Forwarded-For', '')
    if fwd:
        return fwd.split(',')[0].strip()
    return request.remote_addr or ''


@app.route('/status', methods=['GET'])
def status():
    n = count_rows()
    return jsonify({'count': n, 'limit': LIMIT, 'full': n >= LIMIT})


@app.route('/reserve', methods=['POST'])
def reserve():
    payload = request.get_json(silent=True) or {}
    name     = (payload.get('name')     or '').strip()[:200]
    address  = (payload.get('address')  or '').strip()[:1000]
    location = (payload.get('location') or '').strip()[:300]
    lat_raw  = (payload.get('lat')      or '').strip()[:30]
    lon_raw  = (payload.get('lon')      or '').strip()[:30]

    # Coerce lat/lon to floats so Excel sorts and maps them; silently drop on parse error.
    try:    lat = float(lat_raw) if lat_raw else ''
    except: lat = ''
    try:    lon = float(lon_raw) if lon_raw else ''
    except: lon = ''

    if not name or not address:
        return jsonify({'ok': False, 'error': 'Name and address are required'}), 400

    with _lock:
        ensure_workbook()
        wb = load_workbook(XLSX)
        ws = wb.active
        current = ws.max_row - 1
        if current >= LIMIT:
            wb.close()
            return jsonify({'ok': False, 'error': 'full',
                            'count': current, 'limit': LIMIT}), 409

        position = current + 1
        ws.append([
            datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S'),
            position, name, address, location, client_ip(),
            lat, lon,
        ])
        wb.save(XLSX)
        wb.close()

    return jsonify({'ok': True, 'position': position,
                    'count': position, 'limit': LIMIT})


@app.route('/health', methods=['GET'])
def health():
    return jsonify({'ok': True}), 200


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5051)
